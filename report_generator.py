import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared openpyxl style objects, defined once at module level so every
# sheet-writing method below reuses the same look (rather than
# re-instantiating identical Font/Fill/Border objects on every call).
# ---------------------------------------------------------------------------
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, size=11, color="555555")
CAPTION_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
THIN_BORDER = Border(*[Side(style='thin', color='CCCCCC')] * 4)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# Dataset-specific column layouts: (header label, key(s) to look up on the record, dataset(s) it applies to)
# Each dataset has different available fields (e.g. only Dataset C has a
# Price column), so the set and order of columns shown in its Excel sheet
# is defined per dataset here, rather than trying to force all three into
# one identical table shape.
DATASET_COLUMNS = {
    'A': ['eBook ISBN', 'Title', 'Author', 'Edition', 'Year', 'Quantity', 'Recommended by',
          'Relationship', 'Relevance', 'Recency', 'Collection_Support', 'Fuzzy_Score'],
    'B': ['eBook ISBN', 'Title', 'Author', 'Edition', 'Year', 'eBook Format',
          'Discipline (Level 1)', 'Discipline (Level 3)', 'Discipline (Level 4)',
          'Relationship', 'Relevance', 'Recency', 'Format_Suit', 'Fuzzy_Score'],
    'C': ['eISBN', 'Title', 'Author', 'Edition', 'Year', 'eBook Format', 'Price',
          'Category', 'Discipline', 'Relationship', 'Relevance', 'Recency',
          'Format_Suit', 'Affordability', 'Fuzzy_Score'],
}

DATASET_TITLES = {
    'A': 'Dataset A (Existing eBook Collection)',
    'B': 'Dataset B (Academic eBook Catalogue)',
    'C': 'Dataset C (eBook Acquisition / Licensing Catalogue)',
}

SCENARIO_TITLES = {
    1: 'AI, Programming & Mathematical Foundations',
    2: 'Cybersecurity & Secure Computing',
}


class ReportGenerator:
    """
    Turns the in-memory evaluation results (as produced by
    ScenarioEvaluator) into the final Excel workbook deliverable, using
    openpyxl. Responsible for all sheet layout, styling and column
    mapping - none of the analysis logic lives here, only presentation.
    """
    def __init__(self, base_dir=None):
        self.base_dir = base_dir

    # ------------------------------------------------------------------
    # Field lookup: map a display header to the possible dict/row keys
    # ------------------------------------------------------------------
    def get_row_value_by_header(self, row, header, dataset_name):
        """
        Given a display column header (e.g. 'Year') and a record dict,
        finds the actual value to put in that Excel cell.

        Why this indirection is needed: the three datasets don't share a
        common schema (Dataset B calls its year column 'Copyright',
        A and C call it 'Copyright Year'; ISBN fields are named
        differently in each; fuzzy-engine output keys are lowercase while
        the report wants Title_Case headers, etc). Rather than forcing a
        single naming convention everywhere, this method tries a list of
        plausible source-key candidates per header/dataset and returns
        the first one that actually has a usable value.
        """
        if dataset_name == 'B':
            mapping = {
                'eBook ISBN': ['eBook ISBN', 'Print ISBN'],
                'Edition': ['Ed', 'Edition'],
                'Year': ['Copyright'],
                'eBook Format': ['eBook Format'],
            }
        else:  # Dataset A and C
            mapping = {
                'eBook ISBN': ['eBook ISBN', 'eISBN', 'VBID'],
                'eISBN': ['eISBN', 'eBook ISBN', 'VBID'],
                'Edition': ['Edition'],
                'Year': ['Copyright Year'],
            }

        # Headers whose source key is the same across all datasets (plus
        # a couple that fall back to the fuzzy engine's lowercase keys
        # when the Title_Case display key isn't present, e.g. because the
        # record came straight from fuzzy scoring rather than the report
        # layer).
        standard_keys = {
            'Title': ['Title'],
            'Author': ['Author'],
            'Quantity': ['Quantity'],
            'Recommended by': ['Recommended by'],
            'Category': ['Category'],
            'Discipline': ['Discipline'],
            'Discipline (Level 1)': ['Discipline (Level 1)'],
            'Discipline (Level 3)': ['Discipline (Level 3)'],
            'Discipline (Level 4)': ['Discipline (Level 4)'],
            'Relationship': ['Relationship'],
            'Relevance': ['Relevance', 'topic_relevance'],
            'Recency': ['Recency', 'recency_score'],
            'Format_Suit': ['Format_Suit', 'format_suitability'],
            'Affordability': ['Affordability', 'affordability_score'],
            'Collection_Support': ['Collection_Support'],
            'Price': ['Price', 'April List Price (USD)'],
            'Fuzzy_Score': ['Fuzzy_Score', 'fuzzy_score'],
        }

        # Try the dataset-specific mapping first, then fall back to the
        # standard keys (which default to just the header name itself if
        # no explicit mapping exists), taking the first candidate key
        # that's present and not NaN.
        keys_to_try = mapping.get(header, []) + standard_keys.get(header, [header])
        for k in keys_to_try:
            if k in row and row[k] is not None:
                val = row[k]
                if isinstance(val, float) and pd.isna(val):
                    continue
                return val
        return None

    # ------------------------------------------------------------------
    # Sheet-writing helpers
    # ------------------------------------------------------------------
    def _style_header_row(self, ws, row, ncols):
        """Applies the shared header look (bold white text on dark blue
        fill, centered, wrapped, bordered) to every cell in a header row."""
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_table_data(self, ws, header_row, header_labels, df, dataset_name):
        """Writes headers at header_row, then data rows below it. Returns first empty row after the table."""
        ncols = len(header_labels)
        for c, h in enumerate(header_labels, start=1):
            ws.cell(row=header_row, column=c, value=h)
        self._style_header_row(ws, header_row, ncols)

        for r_idx, (_, row) in enumerate(df.iterrows()):
            curr_row = header_row + 1 + r_idx
            for col_idx, h in enumerate(header_labels, start=1):
                if h in ('Predicate_Rank', 'Fuzzy_Rank'):
                    # These two "rank" columns aren't stored data - they're
                    # just the row's position within the already-sorted
                    # DataFrame, so they're computed here rather than
                    # looked up.
                    val = r_idx + 1
                else:
                    val = self.get_row_value_by_header(row, h, dataset_name)
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.border = THIN_BORDER
        return header_row + 1 + len(df)

    def _autosize(self, ws, min_width=10, max_width=45):
        """Roughly auto-sizes each column to fit its longest visible cell
        value (only considering the first line, so wrapped multi-line
        cells don't blow the width out), clamped between min_width and
        max_width so no column becomes unreasonably narrow or wide."""
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value).split("\n")[0])
                col = cell.column_letter
                widths[col] = max(widths.get(col, 0), length)
        for col, w in widths.items():
            ws.column_widths = getattr(ws, 'column_widths', {})
            ws.column_dimensions[col].width = max(min_width, min(max_width, w + 2))

    # ------------------------------------------------------------------
    # Per-dataset scenario sheet
    # ------------------------------------------------------------------
    def _write_dataset_sheet(self, wb, sc_id, dataset_name, sc_results, top_limit):
        """
        Builds one worksheet (e.g. "S1_DatasetA") containing, for a
        single scenario/dataset combination: a predicate-only results
        table (ordered by year, showing every match) followed by a
        fuzzy-enhanced results table (ordered by fuzzy score, limited to
        the scenario's display cap) - directly mirroring the "Predicate
        Query -> Predicate-only Results -> Fuzzy Evaluation ->
        Fuzzy-enhanced Results" pipeline the assessment asks for.
        """
        sheet_name = f"S{sc_id}_Dataset{dataset_name}"
        ws = wb.create_sheet(sheet_name)

        passed_full = sc_results[dataset_name]['passed_full']
        passed_top = sc_results[dataset_name]['passed_top']

        # Predicate-only view: every record that passed, ordered the way
        # a simple non-fuzzy system would (by year, i.e. BaselineRank).
        df_pred = passed_full.sort_values(by='BaselineRank').reset_index(drop=True) if not passed_full.empty else passed_full
        # Fuzzy-enhanced view: the (already limited) top records, ordered
        # by their fuzzy rank instead.
        df_fuzzy = passed_top.sort_values(by='FuzzyRank').reset_index(drop=True) if not passed_top.empty else passed_top

        headers = ['Predicate_Rank'] + DATASET_COLUMNS[dataset_name]
        fuzzy_headers = ['Fuzzy_Rank'] + DATASET_COLUMNS[dataset_name]

        ws.cell(row=1, column=1, value=f"Scenario {sc_id} - {DATASET_TITLES[dataset_name]}").font = TITLE_FONT
        ws.cell(row=2, column=1, value=SCENARIO_TITLES[sc_id]).font = SUBTITLE_FONT
        ws.row_dimensions[1].height = 20

        n_matches = len(passed_full)
        cap1 = f"Predicate-only results ({n_matches} record(s) satisfy the query conditions; ordered by Year, descending)"
        ws.cell(row=4, column=1, value=cap1).font = CAPTION_FONT

        pred_header_row = 6
        next_row = self.write_table_data(ws, pred_header_row, headers, df_pred, dataset_name)

        caption_row = next_row + 1
        # Build a caption describing exactly what's shown in the second
        # (fuzzy) table, since the display rule differs by dataset/
        # scenario: Dataset A in Scenario 2 shows *all* matches (per the
        # assessment's "display all relevant Current Subscription
        # records" rule), everything else is capped at top_limit.
        if dataset_name == 'A':
            desc = "all matches (current subscription), ranked by Fuzzy_Score, descending" if sc_id == 2 \
                else f"all matches, max {top_limit}, ranked by Fuzzy_Score, descending"
        elif n_matches <= top_limit:
            desc = f"all matches (only {n_matches} matches found), ranked by Fuzzy_Score, descending"
        else:
            desc = f"top {top_limit} (only {n_matches} matches found), ranked by Fuzzy_Score, descending"

        cap2 = f"Fuzzy-enhanced results ({desc})"
        ws.cell(row=caption_row, column=1, value=cap2).font = CAPTION_FONT

        fuzzy_header_row = caption_row + 2
        self.write_table_data(ws, fuzzy_header_row, fuzzy_headers, df_fuzzy, dataset_name)

        # Freeze the top rows so the predicate table's header stays
        # visible while scrolling, and give every column a sensible
        # fixed width (with the Title column widened, since titles tend
        # to be the longest text in the sheet).
        ws.freeze_panes = "A7"
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.column_dimensions['C'].width = 42

    # ------------------------------------------------------------------
    # Comparison sheet
    # ------------------------------------------------------------------
    def _write_comparison_sheet(self, wb, sc_id, sc_results, limit_non_a):
        """
        Builds a single cross-dataset "Comparison_S<n>" sheet that lines
        up, for each dataset's top records, the predicate-only rank next
        to the fuzzy rank and how far the record moved - this is the
        direct evidence for the assessment's required "critical
        comparison of predicate-only and fuzzy-enhanced results".
        """
        sheet_name = f"Comparison_S{sc_id}"
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=f"Comparison - Scenario {sc_id} (all three datasets)").font = TITLE_FONT
        ws.cell(row=2, column=1, value="Predicate order (by Year) vs Fuzzy rank (by Fuzzy_Score) - "
                                        "positive Rank_Movement = moved UP after fuzzy evaluation").font = SUBTITLE_FONT

        headers = ['Scenario', 'Dataset', 'Title', 'Predicate_Order_Rank', 'Fuzzy_Rank', 'Fuzzy_Score', 'Rank_Movement']
        header_row = 4
        for c, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=c, value=h)
        self._style_header_row(ws, header_row, len(headers))

        records = []
        for dataset_name in ['A', 'B', 'C']:
            df_passed = sc_results[dataset_name]['passed_full']
            if df_passed.empty:
                continue
            df_passed = df_passed.sort_values(by='BaselineRank')
            # Dataset A only ever contributes 1 row to keep the comparison
            # sheet focused on the more data-rich B/C datasets; the other
            # two use the scenario's normal display limit.
            limit = 1 if dataset_name == 'A' else limit_non_a
            for _, r in df_passed.head(limit).iterrows():
                records.append([
                    f"Scenario {sc_id}", f"Dataset {dataset_name}", r['Title'],
                    r['BaselineRank'], r['FuzzyRank'], r['fuzzy_score'], r['RankChange']
                ])

        for r_idx, rec in enumerate(records):
            curr_row = header_row + 1 + r_idx
            for c, val in enumerate(rec, start=1):
                cell = ws.cell(row=curr_row, column=c, value=val)
                cell.border = THIN_BORDER

        # Explanatory note under the table so a reader doesn't need to
        # infer what "Rank_Movement" means from the numbers alone.
        note_row = header_row + 1 + len(records) + 1
        ws.cell(row=note_row, column=1,
                value="Rank_Movement = Predicate_Order_Rank - Fuzzy_Rank. A positive number means the fuzzy "
                      "system ranked the record higher (more suitable) than plain recency ordering would; a "
                      "negative number means fuzzy reasoning pushed it down (e.g. despite being recent, it "
                      "scored lower on relevance, format or price).").alignment = WRAP_ALIGN

        ws.column_dimensions['C'].width = 55
        for col in ['A', 'B', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 16

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    def _write_summary_sheet(self, wb, sc1_results, sc2_results, dataset_sizes, notes):
        """
        Builds the "Summary_Analysis" sheet: a compact table of match
        counts/coverage per scenario+dataset, followed by the free-text
        discussion notes (generated in main.py) explaining how dataset
        size/structure affected the results - covering the assessment's
        "Comparison and Analysis" requirement in one place.
        """
        ws = wb.create_sheet("Summary_Analysis")
        ws.cell(row=1, column=1, value="Summary & Cross-Dataset Analysis").font = TITLE_FONT
        ws.cell(row=2, column=1, value="Counts and discussion points for the report").font = SUBTITLE_FONT

        ws.cell(row=4, column=1, value="Predicate match counts and coverage by dataset").font = CAPTION_FONT
        headers = ['Scenario', 'Dataset', 'Dataset size (rows)', 'Predicate matches', 'Match rate (%)',
                   'Displayed in fuzzy-ranked output']
        header_row = 5
        for c, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=c, value=h)
        self._style_header_row(ws, header_row, len(headers))

        dataset_labels = {'A': 'A - Existing Collection', 'B': 'B - Academic Catalogue', 'C': 'C - Acquisition/Licensing'}
        row = header_row + 1
        for sc_id, sc_results, limit in [(1, sc1_results, 5), (2, sc2_results, 10)]:
            for dataset_name in ['A', 'B', 'C']:
                size = dataset_sizes[dataset_name]
                matches = len(sc_results[dataset_name]['passed_full'])
                rate = round(matches / size * 100, 2) if size else 0
                # Describe how many of the matches actually made it into
                # the fuzzy-ranked output table shown in the dataset
                # sheet (all of them, all-but-capped, or a capped subset).
                if dataset_name == 'A':
                    disp = f"{matches} (all)"
                elif matches <= limit:
                    disp = f"{matches} (all, {limit} shown max)"
                else:
                    disp = f"{limit} (of {matches})"
                vals = [f"Scenario {sc_id}", dataset_labels[dataset_name], size, matches, rate, disp]
                for c, v in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = THIN_BORDER
                row += 1

        row += 1
        ws.cell(row=row, column=1, value="How dataset size, structure and available evidence affected results").font = CAPTION_FONT
        row += 1
        for note in notes:
            ws.cell(row=row, column=1, value=note).alignment = WRAP_ALIGN
            row += 1

        ws.column_dimensions['A'].width = 100
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20

    # ------------------------------------------------------------------
    # README / Methodology
    # ------------------------------------------------------------------
    def _write_readme(self, wb):
        """Builds the workbook's own in-file "README" sheet: a plain-text
        guide to what each sheet contains, so the workbook is
        self-explanatory even without the accompanying source code."""
        ws = wb.create_sheet("README")
        ws.cell(row=1, column=1, value="BTIS3043 - Non-Written Final Assessment (2026B)").font = TITLE_FONT
        ws.cell(row=2, column=1, value="Predicate + Fuzzy eBook Query System - Output Workbook").font = SUBTITLE_FONT
        lines = [
            "",
            "This workbook is generated automatically by running main.py, which loads the three eBook",
            "datasets, applies predicate queries and fuzzy suitability scoring, and writes the results below.",
            "",
            "Sheet guide:",
            " - Methodology: explains the predicate and fuzzy design used for each dataset.",
            " - S1_DatasetA / S1_DatasetB / S1_DatasetC: Scenario 1 (AI, Programming & Mathematical Foundations)",
            "   results per dataset - predicate-only table, then fuzzy-enhanced table.",
            " - S2_DatasetA / S2_DatasetB / S2_DatasetC: Scenario 2 (Cybersecurity & Secure Computing) results.",
            " - Comparison_S1 / Comparison_S2: side-by-side predicate order vs fuzzy rank for top records.",
            " - Summary_Analysis: match counts/coverage per dataset and discussion notes.",
        ]
        for i, line in enumerate(lines, start=3):
            ws.cell(row=i, column=1, value=line)
        ws.column_dimensions['A'].width = 100

    def _write_methodology(self, wb):
        """Builds the "Methodology" sheet: a plain-language explanation of
        the predicate rules and fuzzy weighting scheme actually
        implemented in predicate_engine.py / fuzzy_engine.py, kept in
        sync with the code by hand (this is documentation, not
        auto-derived from the logic)."""
        ws = wb.create_sheet("Methodology")
        ws.cell(row=1, column=1, value="Methodology").font = TITLE_FONT
        lines = [
            "",
            "Predicate layer:",
            " - Scenario 1: title/discipline keyword matching against AI, programming and mathematics term",
            "   lists, with an explicit exclusion list for unrelated design/vocational titles.",
            " - Scenario 2: title/discipline keyword matching against direct cybersecurity terms, generic",
            "   'security'/'secure' terms qualified by computing context, and a related-terms list (forensics,",
            "   incident response, disaster recovery) that still counts as on-topic but is labelled separately.",
            " - Both scenarios also apply a Year >= 2010 predicate.",
            "",
            "Fuzzy layer:",
            " - Topic relevance: degree of match to the scenario's keyword categories (crisp classification,",
            "   fuzzy-weighted score).",
            " - Recency: linear ramp between 2018 (0) and 2024 (1).",
            " - Format suitability: 1.0 for ePub/PDF, 0.7 for Adobe Reader-only titles, 0.5 otherwise.",
            " - Collection support (Dataset A only): based on quantity already held.",
            " - Affordability (Dataset C only, since it is the only dataset with price data): linear ramp,",
            "   1.0 at <=$100, 0.0 at >=$300.",
            " - The final Fuzzy_Score is a weighted sum of the applicable components (weights differ per",
            "   dataset depending on which attributes are available), rounded to 3 decimal places.",
        ]
        for i, line in enumerate(lines, start=2):
            ws.cell(row=i, column=1, value=line)
        ws.column_dimensions['A'].width = 100

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------
    def generate_excel_results(self, sc1_results, sc2_results, output_dir, dataset_sizes=None, summary_notes=None):
        """
        Public entry point: assembles the entire output workbook (README,
        Methodology, six per-dataset scenario sheets, two comparison
        sheets, and the summary sheet) and saves it to
        <output_dir>/scenario_results.xlsx.
        """
        os.makedirs(output_dir, exist_ok=True)
        unified_file_path = os.path.join(output_dir, "scenario_results.xlsx")

        wb = Workbook()
        wb.remove(wb.active)  # drop default sheet

        self._write_readme(wb)
        self._write_methodology(wb)

        for dataset_name in ['A', 'B', 'C']:
            self._write_dataset_sheet(wb, 1, dataset_name, sc1_results, top_limit=5)
        for dataset_name in ['A', 'B', 'C']:
            self._write_dataset_sheet(wb, 2, dataset_name, sc2_results, top_limit=10)

        self._write_comparison_sheet(wb, 1, sc1_results, limit_non_a=5)
        self._write_comparison_sheet(wb, 2, sc2_results, limit_non_a=10)

        # Both dataset_sizes and summary_notes have safe fallbacks so this
        # method can still produce a workbook even if called without the
        # extra context that main.py normally supplies.
        if dataset_sizes is None:
            dataset_sizes = {
                'A': len(sc1_results['A']['all']),
                'B': len(sc1_results['B']['all']),
                'C': len(sc1_results['C']['all']),
            }
        if summary_notes is None:
            summary_notes = []

        self._write_summary_sheet(wb, sc1_results, sc2_results, dataset_sizes, summary_notes)

        wb.save(unified_file_path)
        print(f"Unified final Excel workbook scenario_results.xlsx successfully created at: {unified_file_path}")
